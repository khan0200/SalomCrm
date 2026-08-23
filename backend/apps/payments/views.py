from decimal import Decimal
import io
from django.http import HttpResponse
from django.db.models import Q
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from apps.core.permissions import IsTenantUser
from apps.students.models import Student
from .models import Payment, PaymentMethodTemplate, PaymentReceiverTemplate, PaymentNotePill
from .serializers import (
    PaymentSerializer, PaymentCreateSerializer, PaymentWithdrawSerializer,
    PaymentEditSerializer, PaymentOverviewStudentSerializer,
    PaymentMethodTemplateSerializer, PaymentReceiverTemplateSerializer, PaymentNotePillSerializer
)
from .services import record_payment, edit_payment, delete_payment

class PaymentViewSet(viewsets.ModelViewSet):
    """
    Payment History and Transactions ViewSet.
    Supports filtering by student ID, method, received_by, date range, and search.
    """
    serializer_class = PaymentSerializer
    permission_classes = [IsTenantUser]

    def get_queryset(self):
        user = self.request.user
        tenant = getattr(self.request, 'tenant', None) or getattr(user, 'tenant', None)

        if user.is_superuser or getattr(user, 'role', '') == 'SUPER_ADMIN':
            tenant_param = self.request.query_params.get('tenant_id')
            if tenant_param:
                qs = Payment.objects.filter(tenant_id=tenant_param)
            elif tenant:
                qs = Payment.objects.filter(tenant=tenant)
            else:
                qs = Payment.objects.all()
        else:
            qs = Payment.objects.filter(tenant=user.tenant)

        # Filters
        student_id = self.request.query_params.get('student_id')
        if student_id:
            qs = qs.filter(student_id=student_id.upper())

        search = self.request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(student_id__icontains=search) |
                Q(student_name__icontains=search) |
                Q(student__full_name__icontains=search) |
                Q(notes__icontains=search) |
                Q(received_by__icontains=search)
            )

        method = self.request.query_params.get('method')
        if method and method != 'all':
            qs = qs.filter(method=method)

        received_by = self.request.query_params.get('received_by')
        if received_by and received_by != 'all':
            qs = qs.filter(received_by=received_by)

        is_discount = self.request.query_params.get('is_discount')
        if is_discount is not None:
            qs = qs.filter(is_discount=(is_discount.lower() == 'true'))

        is_withdrawal = self.request.query_params.get('is_withdrawal')
        if is_withdrawal is not None:
            qs = qs.filter(is_withdrawal=(is_withdrawal.lower() == 'true'))

        return qs.select_related('student', 'created_by').order_by('-created_at')

    def create(self, request, *args, **kwargs):
        serializer = PaymentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        tenant = getattr(request, 'tenant', None) or getattr(user, 'tenant', None)

        student_id = serializer.validated_data.get('student_id')
        student = None
        if student_id:
            student = Student.objects.filter(id=student_id.upper(), tenant=tenant).first()

        payment = record_payment(
            tenant=tenant,
            student=student,
            amount=serializer.validated_data['amount'],
            method=serializer.validated_data['method'],
            received_by=serializer.validated_data['received_by'],
            notes=serializer.validated_data.get('notes') or '',
            is_discount=serializer.validated_data.get('is_discount', False),
            user=user
        )

        return Response(PaymentSerializer(payment).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def withdraw(self, request):
        serializer = PaymentWithdrawSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        tenant = getattr(request, 'tenant', None) or getattr(user, 'tenant', None)

        student_id = serializer.validated_data.get('student_id')
        student = None
        if student_id:
            student = Student.objects.filter(id=student_id.upper(), tenant=tenant).first()

        payment = record_payment(
            tenant=tenant,
            student=student,
            amount=serializer.validated_data['amount'],
            method='Withdrawal',
            received_by='System',
            notes=f"WITHDRAWAL: {serializer.validated_data['reason'].strip().upper()}",
            is_discount=False,
            is_withdrawal=True,
            user=user
        )

        return Response(PaymentSerializer(payment).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        payment = self.get_object()
        serializer = PaymentEditSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        updated_payment = edit_payment(
            payment=payment,
            amount=serializer.validated_data['amount'],
            method=serializer.validated_data['method'],
            received_by=serializer.validated_data['received_by'],
            notes=serializer.validated_data.get('notes'),
            user=request.user
        )
        return Response(PaymentSerializer(updated_payment).data)

    def destroy(self, request, *args, **kwargs):
        payment = self.get_object()
        delete_payment(payment, user=request.user)
        return Response(status=status.HTTP_204_NO_CONTENT)


class PaymentOverviewViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Student Financial Overview ViewSet.
    Calculates and displays students' financial positions (Balance, Discount, Tariff).
    """
    serializer_class = PaymentOverviewStudentSerializer
    permission_classes = [IsTenantUser]

    def get_queryset(self):
        user = self.request.user
        tenant = getattr(self.request, 'tenant', None) or getattr(user, 'tenant', None)

        if user.is_superuser or getattr(user, 'role', '') == 'SUPER_ADMIN':
            qs = Student.objects.filter(tenant=tenant) if tenant else Student.objects.all()
        else:
            qs = Student.objects.filter(tenant=user.tenant)

        # Status filter (active vs archive)
        status_filter = self.request.query_params.get('status', 'Active')
        if status_filter == 'Archive':
            qs = qs.filter(is_deleted=True)
        else:
            qs = qs.filter(is_deleted=False)

        # Search
        search = self.request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(id__icontains=search) |
                Q(full_name__icontains=search) |
                Q(phone1__icontains=search)
            )

        # Tariff filter
        tariffs = self.request.query_params.getlist('tariff') or (self.request.query_params.get('tariff', '').split(',') if self.request.query_params.get('tariff') else [])
        if tariffs and tariffs[0]:
            if 'No Tariff' in tariffs:
                qs = qs.filter(Q(tariff__in=tariffs) | Q(tariff__isnull=True) | Q(tariff=''))
            else:
                qs = qs.filter(tariff__in=tariffs)

        # Group filter
        groups = self.request.query_params.getlist('group') or (self.request.query_params.get('group', '').split(',') if self.request.query_params.get('group') else [])
        if groups and groups[0]:
            qs = qs.filter(student_group__in=groups)

        # Balance Range filter
        balances = self.request.query_params.getlist('balance') or (self.request.query_params.get('balance', '').split(',') if self.request.query_params.get('balance') else [])
        if balances and balances[0]:
            bal_q = Q()
            for b in balances:
                if 'Debt' in b or '< 0' in b:
                    bal_q = bal_q | Q(balance__lt=0)
                elif 'Fully Paid' in b or '= 0' in b:
                    bal_q = bal_q | Q(balance=0)
                elif '> 500,000' in b or '> 500000' in b:
                    bal_q = bal_q | Q(balance__gt=500000)
                elif '> 1,000,000' in b or '> 1000000' in b:
                    bal_q = bal_q | Q(balance__gt=1000000)
                elif '> 2,000,000' in b or '> 2000000' in b:
                    bal_q = bal_q | Q(balance__gt=2000000)
                elif '> 5,000,000' in b or '> 5000000' in b:
                    bal_q = bal_q | Q(balance__gt=5000000)
                elif '> 10,000,000' in b or '> 10000000' in b:
                    bal_q = bal_q | Q(balance__gt=10000000)
            qs = qs.filter(bal_q)

        return qs.order_by('id')


class PaymentExportView(APIView):
    """
    Exports payment history to Excel XLSX spreadsheet.
    """
    permission_classes = [IsTenantUser]

    def get(self, request):
        user = request.user
        tenant = getattr(request, 'tenant', None) or getattr(user, 'tenant', None)

        qs = Payment.objects.filter(tenant=tenant).select_related('student', 'created_by').order_by('-created_at')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Payment History")
        else:
            ws.title = "Payment History"

        # Headers
        headers = ["Date & Time", "Student ID", "Student Name", "Amount (UZS)", "Payment Method", "Received By", "Type", "Notes", "Recorded By"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for p in qs:
            p_type = "Discount" if p.is_discount else ("Withdrawal" if p.is_withdrawal else "Payment")
            ws.append([
                p.created_at.strftime('%Y-%m-%d %H:%M'),
                p.student_id or '-',
                (p.student.full_name if p.student else p.student_name) or '-',
                float(p.amount),
                p.method,
                p.received_by,
                p_type,
                p.notes or '',
                p.created_by.full_name if p.created_by else 'System'
            ])

        # Auto column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_idx = col[0].column
            if col_idx is not None:
                col_letter = get_column_letter(int(col_idx))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="salom_crm_payment_history.xlsx"'
        return response


class BasePaymentOptionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsTenantUser]
    pagination_class = None
    model_class = None

    def get_queryset(self):
        user = self.request.user
        tenant = getattr(self.request, 'tenant', None) or getattr(user, 'tenant', None)
        if not tenant:
            from apps.tenants.models import Tenant
            tenant = Tenant.objects.first()
        if not tenant or self.model_class is None:
            return Payment.objects.none()
        return self.model_class.objects.filter(tenant=tenant)

    def perform_create(self, serializer):
        user = self.request.user
        tenant = getattr(self.request, 'tenant', None) or getattr(user, 'tenant', None)
        if not tenant:
            from apps.tenants.models import Tenant
            tenant = Tenant.objects.first()
        serializer.save(tenant=tenant)


class PaymentMethodTemplateViewSet(BasePaymentOptionViewSet):
    model_class = PaymentMethodTemplate
    serializer_class = PaymentMethodTemplateSerializer


class PaymentReceiverTemplateViewSet(BasePaymentOptionViewSet):
    model_class = PaymentReceiverTemplate
    serializer_class = PaymentReceiverTemplateSerializer


class PaymentNotePillViewSet(BasePaymentOptionViewSet):
    model_class = PaymentNotePill
    serializer_class = PaymentNotePillSerializer


