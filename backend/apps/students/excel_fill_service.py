"""
Excel Fill Service for Salom CRM.
Provides universal parsing, semantic column matching, and high-fidelity template
filling for university Excel (.xlsx) files preserving fonts, colors, borders, and layouts.
"""

import io
import re
import os
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, date
import openpyxl
from copy import copy

logger = logging.getLogger(__name__)

# Uzbekistan regions normalization dictionary
UZB_REGIONS = {
    'TASHKENT': 'TASHKENT', 'TOSHKENT': 'TASHKENT',
    'ANDIJAN': 'ANDIJAN', 'ANDIJON': 'ANDIJAN',
    'SAMARKAND': 'SAMARKAND', 'SAMARQAND': 'SAMARKAND',
    'FERGANA': 'FERGANA', 'FARG\'ONA': 'FERGANA', 'FARGONA': 'FERGANA',
    'NAMANGAN': 'NAMANGAN',
    'BUKHARA': 'BUKHARA', 'BUXORO': 'BUKHARA',
    'KHOREZM': 'KHOREZM', 'XORAZM': 'KHOREZM',
    'KASHKADARYA': 'KASHKADARYA', 'QASHQADARYO': 'KASHKADARYA', 'QASHKADARYO': 'KASHKADARYA',
    'SURKHANDARYA': 'SURKHANDARYA', 'SURXONDARYO': 'SURKHANDARYA',
    'JIZZAKH': 'JIZZAKH', 'JIZZAX': 'JIZZAKH',
    'SIRDARYA': 'SIRDARYA', 'SIRDARYO': 'SIRDARYA',
    'NAVOIY': 'NAVOIY', 'NAVOI': 'NAVOIY',
    'KARAKALPAKSTAN': 'KARAKALPAKSTAN', 'QORAQALPOG\'ISTON': 'KARAKALPAKSTAN', 'QORAQALPOQ': 'KARAKALPAKSTAN'
}

# Uzbekistan district to region mapping
DISTRICT_TO_REGION = {
    # Andijan
    'ANDIJAN': 'ANDIJAN', 'ANDIJON': 'ANDIJAN', 'ASAKA': 'ANDIJAN', 'BALIQCHI': 'ANDIJAN', 'BOZ': 'ANDIJAN', 
    'BULOQBOSHI': 'ANDIJAN', 'IZBOSKAN': 'ANDIJAN', 'JALAKUDUQ': 'ANDIJAN', 'XOJAOBOD': 'ANDIJAN', 
    'QORGONTEPA': 'ANDIJAN', 'MARHAMAT': 'ANDIJAN', 'MARKHAMAT': 'ANDIJAN', 'OLTINKOL': 'ANDIJAN', 
    'PAXTAOBOD': 'ANDIJAN', 'PAKHTABAD': 'ANDIJAN', 'PAKTABAD': 'ANDIJAN', 'SHAHRIXON': 'ANDIJAN', 'ULUGNOR': 'ANDIJAN', 'XONOBOD': 'ANDIJAN',
    # Tashkent
    'TASHKENT': 'TASHKENT', 'TOSHKENT': 'TASHKENT', 'BEKOBOD': 'TASHKENT', 'BOSTANLYK': 'TASHKENT', 'BOSTONLIQ': 'TASHKENT',
    'BUKA': 'TASHKENT', 'CHINOZ': 'TASHKENT', 'QIBRAY': 'TASHKENT', 'KIBRAY': 'TASHKENT', 
    'PARKENT': 'TASHKENT', 'PISKENT': 'TASHKENT', 'QUYICHIRCHIQ': 'TASHKENT', 'ORTACHIRCHIQ': 'TASHKENT',
    'YUQORICHIRCHIQ': 'TASHKENT', 'YANGIYOL': 'TASHKENT', 'ZANGIOTA': 'TASHKENT', 'CHILANZAR': 'TASHKENT', 
    'YUNUSABAD': 'TASHKENT', 'MIRZO ULUGBEK': 'TASHKENT', 'YASHNABOD': 'TASHKENT', 'SHAYXONTOHUR': 'TASHKENT',
    # Samarkand
    'SAMARKAND': 'SAMARKAND', 'SAMARQAND': 'SAMARKAND', 'BULUNGUR': 'SAMARKAND', 'ISHTIXON': 'SAMARKAND', 
    'JOMBOY': 'SAMARKAND', 'KATTAQORGON': 'SAMARKAND', 'QOSHRABOT': 'SAMARKAND', 'NARPAY': 'SAMARKAND', 
    'NUROBOD': 'SAMARKAND', 'OQDARYO': 'SAMARKAND', 'PASTDARGOM': 'SAMARKAND', 'PAXTACHI': 'SAMARKAND', 
    'PAYARIQ': 'SAMARKAND', 'TOYLOQ': 'SAMARKAND', 'URGUT': 'SAMARKAND',
    # Fergana
    'FERGANA': 'FERGANA', 'FARG\'ONA': 'FERGANA', 'BESHARIK': 'FERGANA', 'BOGDOD': 'FERGANA', 'BUVAYDA': 'FERGANA', 
    'DANGARA': 'FERGANA', 'FARGONA': 'FERGANA', 'FURQAT': 'FERGANA', 'QOQON': 'FERGANA', 'KOKAND': 'FERGANA', 
    'QUVA': 'FERGANA', 'QUVASOY': 'FERGANA', 'RISHTON': 'FERGANA', 'SOX': 'FERGANA', 'TOSHLOQ': 'FERGANA', 
    'UCHKOPRIK': 'FERGANA', 'YOZYOVON': 'FERGANA', 'MARGILON': 'FERGANA',
    # Namangan
    'NAMANGAN': 'NAMANGAN', 'CHORTOQ': 'NAMANGAN', 'CHUST': 'NAMANGAN', 'KOSONSOY': 'NAMANGAN', 
    'MINGBULOQ': 'NAMANGAN', 'NORIN': 'NAMANGAN', 'POP': 'NAMANGAN', 'TORAQORGON': 'NAMANGAN', 
    'UCHQORGON': 'NAMANGAN', 'UYCHI': 'NAMANGAN', 'YANGIQORGON': 'NAMANGAN',
    # Bukhara
    'BUKHARA': 'BUKHARA', 'BUXORO': 'BUKHARA', 'G`IJDUVON': 'BUKHARA', 'GIJDUVAN': 'BUKHARA', 'JONDOR': 'BUKHARA', 
    'KOGON': 'BUKHARA', 'QORAKOL': 'BUKHARA', 'QOROVULBOZOR': 'BUKHARA', 'OLOT': 'BUKHARA', 'PESHKU': 'BUKHARA', 
    'ROMITAN': 'BUKHARA', 'SHOFIRKON': 'BUKHARA', 'VOBKENT': 'BUKHARA',
    # Sirdarya
    'SIRDARYA': 'SIRDARYA', 'SIRDARYO': 'SIRDARYA', 'BOYOVUT': 'SIRDARYA', 'GULISTON': 'SIRDARYA', 
    'XOVOS': 'SIRDARYA', 'MEHNATOBOD': 'SIRDARYA', 'MIRZAOBOD': 'SIRDARYA', 'OQOLTIN': 'SIRDARYA', 
    'SARDABA': 'SIRDARYA', 'SARDOBA': 'SIRDARYA', 'SAYXUNOBOD': 'SIRDARYA', 'YANGIYER': 'SIRDARYA',
    # Kashkadarya
    'KASHKADARYA': 'KASHKADARYA', 'QASHQADARYO': 'KASHKADARYA', 'CHIROQCHI': 'KASHKADARYA', 'DEHQONOBOD': 'KASHKADARYA', 
    'GUZOR': 'KASHKADARYA', 'QAMASHI': 'KASHKADARYA', 'QARSHI': 'KASHKADARYA', 'KASBI': 'KASHKADARYA', 
    'KITOB': 'KASHKADARYA', 'KOSON': 'KASHKADARYA', 'MIRISHKOR': 'KASHKADARYA', 'MUBORAK': 'KASHKADARYA', 
    'NISHON': 'KASHKADARYA', 'SHAHRISABZ': 'KASHKADARYA', 'YAKKABOG': 'KASHKADARYA',
    # Surkhandarya
    'SURKHANDARYA': 'SURKHANDARYA', 'SURXONDARYO': 'SURKHANDARYA', 'ANGOR': 'SURKHANDARYA', 'BANDIXON': 'SURKHANDARYA', 
    'BOYSUN': 'SURKHANDARYA', 'DENOV': 'SURKHANDARYA', 'JARQORGON': 'SURKHANDARYA', 'QIZIRIQ': 'SURKHANDARYA', 
    'QUMQORGON': 'SURKHANDARYA', 'MUZRABOT': 'SURKHANDARYA', 'OLTINSOY': 'SURKHANDARYA', 'SARIOSIYO': 'SURKHANDARYA', 
    'SHEROBOD': 'SURKHANDARYA', 'SHO`RCHI': 'SURKHANDARYA', 'TERMIZ': 'SURKHANDARYA', 'UZUN': 'SURKHANDARYA',
    # Khorezm
    'KHOREZM': 'KHOREZM', 'XORAZM': 'KHOREZM', 'BOGOT': 'KHOREZM', 'GURLAN': 'KHOREZM', 'XIVA': 'KHOREZM', 
    'XAZORASP': 'KHOREZM', 'XONQA': 'KHOREZM', 'QOSHKOPIR': 'KHOREZM', 'SHOVOT': 'KHOREZM', 'URGANCH': 'KHOREZM', 
    'YANGIARIQ': 'KHOREZM', 'YANGIBOZOR': 'KHOREZM',
    # Jizzakh
    'JIZZAKH': 'JIZZAKH', 'JIZZAX': 'JIZZAKH', 'ARNASOY': 'JIZZAKH', 'BAXMAL': 'JIZZAKH', 'DOSTLIK': 'JIZZAKH', 
    'FORISH': 'JIZZAKH', 'GALLAOROL': 'JIZZAKH', 'SHAROF RASHIDOV': 'JIZZAKH', 'MIRZACHOL': 'JIZZAKH', 
    'PAXTAKOR': 'JIZZAKH', 'YANGIOBOD': 'JIZZAKH', 'ZOMIN': 'JIZZAKH', 'ZAFAROBOD': 'JIZZAKH',
    # Navoiy
    'NAVOIY': 'NAVOIY', 'NAVOI': 'NAVOIY', 'KONIMEX': 'NAVOIY', 'QIZILTEPA': 'NAVOIY', 'XATIRCHI': 'NAVOIY', 
    'KARMANA': 'NAVOIY', 'NUROTA': 'NAVOIY', 'TOMDI': 'NAVOIY', 'UCHQUDUQ': 'NAVOIY', 'ZARAFSHON': 'NAVOIY',
    # Karakalpakstan
    'KARAKALPAKSTAN': 'KARAKALPAKSTAN', 'QORAQALPOG\'ISTON': 'KARAKALPAKSTAN', 'NUKUS': 'KARAKALPAKSTAN', 
    'AMUDARYO': 'KARAKALPAKSTAN', 'BERUNIY': 'KARAKALPAKSTAN', 'CHIMBOY': 'KARAKALPAKSTAN', 'ELLIKQALA': 'KARAKALPAKSTAN', 
    'QONLIKO`L': 'KARAKALPAKSTAN', 'QORAO`ZAK': 'KARAKALPAKSTAN', 'MOYNOQ': 'KARAKALPAKSTAN', 
    'SHUMANAY': 'KARAKALPAKSTAN', 'TAXTIKO`PIR': 'KARAKALPAKSTAN', 'TO`RTKO`L': 'KARAKALPAKSTAN', 'XO`JAYLI': 'KARAKALPAKSTAN'
}


def extract_city_and_state(address: str) -> Tuple[str, str]:
    """
    Intelligently extracts City / District and State / Region from a raw full address string.
    """
    if not address or not str(address).strip():
        return '', ''
    
    clean = ' '.join(str(address).upper().replace(',', ' , ').replace('.', ' . ').split())
    
    # 1. State / Region detection directly from text
    detected_state = ''
    for key, val in UZB_REGIONS.items():
        if re.search(r'\b' + re.escape(key) + r'(?:\s+(?:REGION|VILOYATI|VILOYAT|OBLAST))?\b', clean):
            detected_state = val
            break
            
    # 2. City / District extraction
    m = re.search(r'(?:,\s*|\b)([A-Z\'-]+)\s+(?:DISTRICT|DISRTIKT|TUMAN|TUMANI|SHAHAR|SHAHRI|CITY)\b', clean)
    detected_city = ''
    if m:
        word = m.group(1).strip()
        if word not in ['REGION', 'VILOYATI', 'VILOYAT', 'OBLAST', 'STREET', 'ST', 'HOUSE', 'VILLAGE', 'MFY', 'UZBEKISTAN']:
            detected_city = word
            
    # If not found via regex, search for known district names in address
    if not detected_city:
        for dist_key in DISTRICT_TO_REGION.keys():
            if re.search(r'\b' + re.escape(dist_key) + r'\b', clean):
                detected_city = dist_key
                if not detected_state:
                    detected_state = DISTRICT_TO_REGION[dist_key]
                break

    # If state not found yet, infer from detected city
    if detected_city and not detected_state:
        detected_state = DISTRICT_TO_REGION.get(detected_city, '')

    if detected_state:
        detected_state = UZB_REGIONS.get(detected_state, detected_state)

    return detected_city, detected_state


# CRM Field dictionary for semantic mapping and UI selection
CRM_FIELDS = [
    {
        "key": "_sequence_no",
        "label": "Tartib raqami (№ / 순번 / No)",
        "category": "system",
        "description": "Avtomatik 1, 2, 3... tartib raqami qo'yiladi"
    },
    {
        "key": "_static_value",
        "label": "Statik qiymat (Barchaga bir xil matn)",
        "category": "system",
        "description": "Masalan: Agentlik nomi, E-visa, Round 1"
    },
    {
        "key": "_skip",
        "label": "O'tkazib yuborish (Bo'sh qoldirish)",
        "category": "system",
        "description": "Ushbu ustun to'ldirilmaydi"
    },
    # Personal info
    {"key": "full_name", "label": "Full Name / F.I.SH (Inglizcha)", "category": "personal"},
    {"key": "first_name", "label": "Ism (First Name)", "category": "personal"},
    {"key": "last_name", "label": "Familiya (Last Name)", "category": "personal"},
    {"key": "korean_name", "label": "Koreyscha ism (국문이름 / 한글성명)", "category": "personal"},
    {"key": "gender", "label": "Jinsi (Sex / Gender / 성별)", "category": "personal"},
    {"key": "birthday", "label": "Tug'ilgan sana (Birth Date / 생년월일)", "category": "personal"},
    {"key": "nationality", "label": "Fuqaroligi (Nationality / 국적)", "category": "personal"},
    {"key": "address", "label": "To'liq yashash manzili (Full Address / 주소)", "category": "personal"},
    {"key": "address_city", "label": "Shahar / Tuman (City / District / 시·군·구)", "category": "personal"},
    {"key": "address_state", "label": "Viloyat / Region (State / Province / 도·시)", "category": "personal"},
    # Passport
    {"key": "passport", "label": "Pasport seriya va raqam (Passport No / 여권번호)", "category": "passport"},
    {"key": "passport_issue_date", "label": "Pasport berilgan sana (Issue Date / 발급일)", "category": "passport"},
    {"key": "passport_expire_date", "label": "Pasport amal qilish muddati (Expiry Date / 만료일)", "category": "passport"},
    # Contacts
    {"key": "phone1", "label": "Asosiy telefon raqam (Phone 1 / 연락처)", "category": "contacts"},
    {"key": "phone2", "label": "Qo'shimcha telefon (Phone 2 / 비상연락처)", "category": "contacts"},
    {"key": "email", "label": "Email manzil (Email / 이메일)", "category": "contacts"},
    # Parents
    {"key": "father_name", "label": "Otasining ismi (Father's Name / 부 성명)", "category": "parents"},
    {"key": "father_phone", "label": "Otasining telefoni (Father's Phone / 부 연락처)", "category": "parents"},
    {"key": "father_job", "label": "Otasining ish joyi (Father's Job / 부 직업)", "category": "parents"},
    {"key": "mother_name", "label": "Onasining ismi (Mother's Name / 모 성명)", "category": "parents"},
    {"key": "mother_phone", "label": "Onasining telefoni (Mother's Phone / 모 연락처)", "category": "parents"},
    {"key": "mother_job", "label": "Onasining ish joyi (Mother's Job / 모 직업)", "category": "parents"},
    # Education & Certificates
    {"key": "level", "label": "Ta'lim bosqichi (Level / Degree / 과정)", "category": "education"},
    {"key": "major", "label": "Yo'nalish / Mutaxassislik (Major / 전공)", "category": "education"},
    {"key": "final_school_name", "label": "Tugatgan maktab/litsey/universitet (Previous School / University / 최종학력)", "category": "education"},
    {"key": "gpa", "label": "GPA / O'rtacha baho", "category": "education"},
    {"key": "language_certificate", "label": "Til sertifikati nomi (TOPIK / IELTS / SKA)", "category": "education"},
    {"key": "certificate_score", "label": "Sertifikat bali (Score / 급수)", "category": "education"},
    {"key": "certificate_valid_date", "label": "Sertifikat amal qilish muddati", "category": "education"},
]

# Multi-lingual dictionary for automatic matching
SEMANTIC_PATTERNS = [
    # Sequence / No
    (r"^(no|№|n|순번|번호|num|seq|순서)$", "_sequence_no"),
    
    # Telegram / Username (Always Skip)
    (r"(telegram|username|tg_username|telegram_username|telegr)", "_skip"),
    
    # Parents specific FIRST
    (r"(father['’]?s\s*name|father\s*name|father\s*fullname|부\s*성명|부친\s*성명|otasining\s*ismi)", "father_name"),
    (r"(father['’]?s\s*(number|phone|mobile)|father\s*(mobile|phone|no)|부\s*(연락처|전화번호)|부친\s*연락처|otasining\s*telefoni)", "father_phone"),
    (r"(father['’]?s\s*job|부\s*직업|부친\s*직업|otasining\s*kasbi)", "father_job"),
    
    (r"(mother['’]?s\s*name|mother\s*name|mother\s*fullname|모\s*성명|모친\s*성명|onasining\s*ismi)", "mother_name"),
    (r"(mother['’]?s\s*(number|phone|mobile)|mother\s*(mobile|phone|no)|모\s*(연락처|전화번호)|모친\s*연락처|onasining\s*telefoni)", "mother_phone"),
    (r"(mother['’]?s\s*job|모\s*직업|모친\s*직업|onasining\s*kasbi)", "mother_job"),

    # Korean Name
    (r"(korean\s*name|국문\s*이름|국문성명|한글\s*이름|한글성명|korean_name)", "korean_name"),

    # English Full Name / Name (Excludes username)
    (r"(english\s*name|student\s*name|성명\s*\(?영문\)?|영문\s*이름|영문성명|name\s*\(as\s*appreaed|full\s*name|^name$|^f\.?i\.?sh$|^fio$|talaba\s*ismi|фио)", "full_name"),
    
    # Passport Number
    (r"(passport\s*number|pasport\s*number|여권\s*번호|여권|passport\s*no|^passport$|^pasport$|серия\s*и\s*номер\s*паспорта)", "passport"),
    
    # Passport Dates
    (r"(passport\s*issue|issue\s*date|여권\s*발급일|발급일자|date\s*of\s*issue)", "passport_issue_date"),
    (r"(passport\s*expir|expiry\s*date|expiration|여권\s*만료일|만료일자|date\s*of\s*expiration)", "passport_expire_date"),
    
    # Birthday / DOB
    (r"(birth\s*date|date\s*of\s*birth|birthday|생년월일|생일|생년|tugilgan\s*sana|дата\s*рождения|^dob$)", "birthday"),
    
    # Gender / Sex
    (r"(gender|sex|성별|jinsi|пол)", "gender"),
    
    # Nationality
    (r"(nationality|citizenship|국적|fuqaroligi|гражданство)", "nationality"),
    
    # City / District
    (r"^(city|town|district|tuman|shahar|도시|시|군|구)$", "address_city"),
    
    # State / Region / Province
    (r"^(state|province|region|viloyat|도|주)$", "address_state"),
    
    # Phone numbers
    (r"(student\s*number|student\s*phone|phone\s*1|^phone$|연락처|본인\s*연락처|전화번호|telefon|телефон)", "phone1"),
    (r"(phone\s*2|extra\s*phone|비상\s*연락처|비상연락처|qo'shimcha\s*telefon)", "phone2"),
    
    # Email
    (r"(email|e-mail|이메일|elektron\s*pochta)", "email"),
    
    # Address Full
    (r"(address\s*in\s*english|full\s*address|address|주소|manzil|адрес)", "address"),
    
    # Language certificate
    (r"(language\s*certificate|토픽\s*/?\s*ielts|topik|ielts|어학\s*능력|til\s*sertifikati|сертификат)", "language_certificate"),
    (r"(certificate\s*score|topik\s*score|ielts\s*score|급수|점수|sertifikat\s*bali)", "certificate_score"),
    
    # Major / Department
    (r"(major|department|학과명|전공|세부전공|yo'nalish|mutaxassislik|факультет|специальность)", "major"),
    
    # Degree / Course Level
    (r"(course|degree|entry\s*level|전형\s*과정명|학위\s*과정|daraja|bosqich)", "level"),
    
    # School / University (Always Graduated / Previous School in CRM)
    (r"(university|previous\s*university|school|previous\s*school|graduated\s*school|최종\s*학력|출신\s*학교|maktab|litsey|universitet|college)", "final_school_name"),
]


def detect_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, max_scan_rows: int = 15) -> int:
    """
    Scans the top rows of a worksheet to identify which row is the table header.
    Looks for the row with the most non-empty text cells and standard header keywords.
    """
    best_row = 1
    best_score = -1
    
    for r in range(1, min(ws.max_row + 1, max_scan_rows + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 100))]
        non_empty = [str(v).strip() for v in row_vals if v is not None and str(v).strip()]
        
        if not non_empty:
            continue
            
        score = len(non_empty)
        
        for val in non_empty:
            low = val.lower()
            if any(kw in low for kw in ['name', 'no', '№', 'passport', 'birth', 'gender', 'sex', 'city', 'state', '성명', '여권', '순번', '생년', '연락처', '주소']):
                score += 5
            if len(non_empty) <= 2 and len(val) > 30:
                score -= 10
                
        if score > best_score:
            best_score = score
            best_row = r
            
    return best_row


def match_column_field(header_text: str) -> Tuple[str, float]:
    """
    Matches header text to a CRM field key.
    Returns (field_key, confidence_0_to_1).
    """
    if not header_text or not str(header_text).strip():
        return ("_skip", 0.0)
        
    cleaned = " ".join(str(header_text).strip().split()).lower()
    
    # Explicit Telegram / Username check
    if 'telegram' in cleaned or 'username' in cleaned:
        return ("_skip", 0.99)
        
    for pattern, field_key in SEMANTIC_PATTERNS:
        if re.search(pattern, cleaned, re.IGNORECASE):
            return (field_key, 0.95)
            
    # Fuzzy keyword check
    if cleaned in ('city', 'district', 'town', 'tuman', 'shahar'):
        return ('address_city', 0.95)
    if cleaned in ('state', 'province', 'region', 'viloyat'):
        return ('address_state', 0.95)
    if 'name' in cleaned or 'ism' in cleaned or '성명' in cleaned:
        if 'kor' in cleaned or '국문' in cleaned or '한글' in cleaned:
            return ('korean_name', 0.85)
        return ('full_name', 0.85)
    if 'pass' in cleaned or '여권' in cleaned:
        return ('passport', 0.85)
    if 'birth' in cleaned or '생년' in cleaned or 'dob' in cleaned:
        return ('birthday', 0.85)
    if 'phone' in cleaned or 'tel' in cleaned or '연락처' in cleaned:
        if 'father' in cleaned or '부' in cleaned:
            return ('father_phone', 0.85)
        if 'mother' in cleaned or '모' in cleaned:
            return ('mother_phone', 0.85)
        return ('phone1', 0.85)
    if 'mail' in cleaned or '이메일' in cleaned:
        return ('email', 0.85)
    if 'addr' in cleaned or '주소' in cleaned:
        return ('address', 0.85)
    if 'sex' in cleaned or 'gender' in cleaned or '성별' in cleaned:
        return ('gender', 0.85)
    if 'no' in cleaned or '№' in cleaned or '순번' in cleaned:
        return ('_sequence_no', 0.85)
    if 'university' in cleaned or 'school' in cleaned or 'college' in cleaned:
        return ('final_school_name', 0.85)
        
    return ("_skip", 0.1)


def analyze_excel_file(file_bytes: bytes) -> Dict[str, Any]:
    """
    Analyzes an uploaded Excel template file.
    Returns sheets info, detected headers, preview rows, and suggested mappings.
    Only includes columns that actually have a non-empty header name.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets_info = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = detect_header_row(ws)
        
        # Read preview rows (up to 12 rows)
        preview_rows = []
        for r in range(1, min(ws.max_row + 1, 15)):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 60)):
                val = ws.cell(r, c).value
                if isinstance(val, (datetime, date)):
                    val = val.strftime('%Y-%m-%d')
                row_data.append(str(val) if val is not None else "")
            if any(row_data):
                preview_rows.append({"row_idx": r, "values": row_data})
                
        # Analyze columns on header row
        columns = []
        for c in range(1, min(ws.max_column + 1, 80)):
            col_letter = openpyxl.utils.get_column_letter(c)
            header_val = ws.cell(header_row, c).value
            header_str = str(header_val).strip() if header_val is not None else ""
            
            # Skip completely empty header columns
            if not header_str:
                continue
            
            # Check if column is hidden
            col_dim = ws.column_dimensions.get(col_letter)
            is_hidden = getattr(col_dim, 'hidden', False) if col_dim else False
            
            matched_field, confidence = match_column_field(header_str)
            
            # Look at sample value from row below header (if exists)
            sample_val = ""
            if header_row + 1 <= ws.max_row:
                sv = ws.cell(header_row + 1, c).value
                if sv is not None:
                    sample_val = str(sv).strip()
            
            columns.append({
                "col_idx": c,
                "col_letter": col_letter,
                "header_name": header_str,
                "is_hidden": bool(is_hidden),
                "suggested_field": matched_field,
                "confidence": confidence,
                "sample_value": sample_val
            })
            
        sheets_info.append({
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "detected_header_row": header_row,
            "columns": columns,
            "preview_rows": preview_rows
        })
        
    return {
        "sheets": sheets_info,
        "available_fields": CRM_FIELDS
    }


def format_cell_value(
    raw_val: Any,
    field_key: str,
    format_rules: Dict[str, Any],
    fallback: str = ""
) -> str:
    """
    Transforms CRM raw data according to column formatting rules (Date, Gender, Phone, Fallbacks).
    """
    if raw_val is None or str(raw_val).strip() == "" or str(raw_val).strip() == "None":
        return fallback

    val_str = str(raw_val).strip()

    # 1. Date formatting
    if field_key in ('birthday', 'passport_issue_date', 'passport_expire_date', 'certificate_valid_date'):
        date_format = format_rules.get('dateFormat', 'YYYY-MM-DD')
        parsed_date = None
        for fmt in ('%Y-%m-%d', '%Y.%m.%d', '%Y%m%d', '%d.%m.%Y', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S'):
            try:
                parsed_date = datetime.strptime(val_str.split()[0], fmt)
                break
            except Exception:
                pass
        
        if parsed_date:
            if date_format == 'YYYY.MM.DD':
                return parsed_date.strftime('%Y.%m.%d')
            elif date_format == 'YYYYMMDD':
                return parsed_date.strftime('%Y%m%d')
            elif date_format == 'DD.MM.YYYY':
                return parsed_date.strftime('%d.%m.%Y')
            elif date_format == 'DD-MM-YYYY':
                return parsed_date.strftime('%d-%m-%Y')
            else:
                return parsed_date.strftime('%Y-%m-%d')
        return val_str

    # 2. Gender formatting
    if field_key == 'gender':
        gender_format = format_rules.get('genderFormat', 'MALE/FEMALE')
        upper_val = val_str.upper()
        # Female is tested first and with whole-word matches: a substring test for
        # 'M' would treat FEMALE (and AYOL's 'F') as male.
        is_female = (
            upper_val.startswith('F')
            or upper_val.startswith('AYOL')
            or upper_val.startswith('W')
            or '여' in upper_val
            or 'ЖЕН' in upper_val
        )
        is_male = not is_female and (
            upper_val.startswith('M')
            or upper_val.startswith('ER')
            or '남' in upper_val
            or 'МУЖ' in upper_val
        )

        if is_male:
            if gender_format == '남/여':
                return '남'
            elif gender_format == '남성/여성':
                return '남성'
            elif gender_format == 'Male/Female':
                return 'Male'
            elif gender_format == 'M/F':
                return 'M'
            else:
                return 'MALE'
        elif is_female:
            if gender_format == '남/여':
                return '여'
            elif gender_format == '남성/여성':
                return '여성'
            elif gender_format == 'Male/Female':
                return 'Female'
            elif gender_format == 'M/F':
                return 'F'
            else:
                return 'FEMALE'
        return val_str

    # 3. Phone formatting
    if 'phone' in field_key:
        phone_format = format_rules.get('phoneFormat', 'original')
        digits = re.sub(r'\D', '', val_str)
        if phone_format == 'dashed' and len(digits) >= 9:
            if len(digits) == 9:
                return f"{digits[0:2]}-{digits[2:5]}-{digits[5:7]}-{digits[7:9]}"
            elif len(digits) == 12:
                return f"+{digits[0:3]} {digits[3:5]}-{digits[5:8]}-{digits[8:10]}-{digits[10:12]}"
        elif phone_format == 'plus_998' and len(digits) == 9:
            return f"+998{digits}"
        elif phone_format == 'digits_only':
            return digits

    return val_str


def generate_filled_excel(
    file_bytes: bytes,
    sheet_name: str,
    column_mappings: List[Dict[str, Any]],
    students_data: List[Dict[str, Any]],
    fill_mode: str = "append",
    start_row: Optional[int] = None,
    auto_increment_sequence: bool = True
) -> io.BytesIO:
    """
    Fills student data into the original Excel template using openpyxl while preserving
    all styling, borders, fonts, colors, merges, and layouts.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    
    detected_header_row = detect_header_row(ws)
    
    # Determine the starting row for writing
    if start_row and start_row > 0:
        target_start_row = start_row
    elif fill_mode == "append":
        # Find first empty row after headers
        last_filled = detected_header_row
        for r in range(detected_header_row + 1, ws.max_row + 1):
            row_has_data = any(
                ws.cell(r, c).value is not None and str(ws.cell(r, c).value).strip() != ""
                for c in range(1, min(ws.max_column + 1, 40))
            )
            if row_has_data:
                last_filled = r
        target_start_row = last_filled + 1
    else:
        # Overwrite mode: start immediately on the row after header
        target_start_row = detected_header_row + 1

    # Find maximum existing sequence number if appending
    current_seq = 0
    if auto_increment_sequence:
        seq_col = None
        for m in column_mappings:
            if m.get('field') == '_sequence_no':
                seq_col = m.get('col_idx')
                break
        if seq_col:
            for r in range(detected_header_row + 1, target_start_row):
                v = ws.cell(r, seq_col).value
                if v is not None and str(v).strip().isdigit():
                    current_seq = max(current_seq, int(str(v).strip()))

    # Determine reference style row (use row above start_row or header row)
    ref_style_row = max(1, target_start_row - 1)

    # Write each student row
    current_row = target_start_row
    for student in students_data:
        current_seq += 1
        
        for mapping in column_mappings:
            col_idx = mapping.get('col_idx')
            field_key = mapping.get('field', '_skip')
            static_val = mapping.get('static_value', '')
            fallback_val = mapping.get('fallback', '')
            format_rules = mapping.get('format_rules', {})
            
            if not col_idx or field_key == '_skip':
                continue
                
            cell = ws.cell(row=current_row, column=col_idx)
            
            # Clone styles from reference row
            ref_cell = ws.cell(row=ref_style_row, column=col_idx)
            if ref_cell.has_style:
                try:
                    cell.font = copy(ref_cell.font)
                    cell.border = copy(ref_cell.border)
                    cell.fill = copy(ref_cell.fill)
                    cell.number_format = copy(ref_cell.number_format)
                    cell.protection = copy(ref_cell.protection)
                    cell.alignment = copy(ref_cell.alignment)
                except Exception as e:
                    logger.debug(f"Style copy error on cell ({current_row}, {col_idx}): {e}")

            # Compute value to write
            if field_key == '_sequence_no':
                cell.value = current_seq
            elif field_key == '_static_value':
                cell.value = static_val
            elif field_key in ('first_name', 'last_name'):
                full = student.get('full_name', '')
                parts = full.split() if full else []
                if field_key == 'first_name':
                    cell.value = parts[0] if parts else fallback_val
                else:
                    cell.value = " ".join(parts[1:]) if len(parts) > 1 else fallback_val
            elif field_key == 'nationality':
                cell.value = student.get('nationality') or 'UZBEKISTAN'
            elif field_key == 'address_city':
                city, _ = extract_city_and_state(student.get('address') or '')
                cell.value = city if city else fallback_val
            elif field_key == 'address_state':
                _, state = extract_city_and_state(student.get('address') or '')
                cell.value = state if state else fallback_val
            else:
                raw_val = student.get(field_key)
                formatted = format_cell_value(raw_val, field_key, format_rules, fallback_val)
                cell.value = formatted

        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
