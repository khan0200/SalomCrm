import re
import io
import uuid
import logging
from typing import Any, cast
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from rest_framework import viewsets, permissions, status
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.request import Request
from django.db.models import Q
from apps.core.permissions import IsTenantUser, IsTenantHeadManager
from .models import (
    Student, Folder, TariffOption, EducationLevelOption,
    StudentGroupOption, LeadSourceOption, CoordinatorOption,
    UniversityOption, UniversityStatusOption, TagOption, SchoolDirectory, MajorOption
)
from .serializers import (
    StudentListSerializer, StudentDetailSerializer, StudentCreateUpdateSerializer,
    StudentSetColorSerializer, StudentSetFoldersSerializer, FolderSerializer,
    TariffOptionSerializer, EducationLevelOptionSerializer, StudentGroupOptionSerializer,
    LeadSourceOptionSerializer, CoordinatorOptionSerializer,
    UniversityOptionSerializer, UniversityStatusOptionSerializer,
    TagOptionSerializer, SchoolDirectorySerializer, MajorOptionSerializer
)
from .services import archive_student, restore_student, permanent_delete_student

logger = logging.getLogger(__name__)

DEFAULT_TAGS_DATA = [
    ('HAL', '✅'),
    ('JEONJU REG', '📋'),
    ('KDB', '💳'),
    ('Natija kutilmoqda', '⏳'),
    ('Topik 2', '🏷️'),
    ('til kursi', '🏷️'),
    ('BUFS TIL KURSI', '🚩'),
    ('BUFS APPFEE', '🎫'),
    ('AeroSpace', '✈️'),
    ('GIMCHEON OK', '🏷️'),
    ('WOOSUK APPFEE', '💳'),
    ('Documents Pending', '📄'),
    ('Visa Processing', '🎫'),
    ('Visa Approved', '🛂'),
    ('Departure', '✈️'),
    ('Arrived', '📍'),
    ('Scholarship Awarded', '💎'),
    ('Call', '📞'),
    ('Apply', '🎓'),
    ('Documents', '📄'),
    ('Payment', '💰'),
]

def alphanumeric_key(student_id):
    """
    Sort key for alphanumeric student IDs matching Uniapp v2 logic.
    Splits into prefix and numeric components (e.g. 'UB120' -> ('UB', 120)).
    """
    if not student_id:
        return ("", 0)
    match = re.match(r'^([A-Za-z\s_-]*)(\d*)$', str(student_id).strip())
    if match:
        prefix = match.group(1).upper()
        num = int(match.group(2)) if match.group(2) else 0
        return (prefix, num)
    return (str(student_id).upper(), 0)


class StudentViewSet(viewsets.ModelViewSet):
    """
    Enterprise Student Management ViewSet with server-side filtering,
    alphanumeric ID sorting, folder scopes, soft archive/restore, and color tagging.
    """
    permission_classes = [IsTenantUser]

    def get_serializer_class(self):
        if self.action == 'list':
            return StudentListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return StudentCreateUpdateSerializer
        return StudentDetailSerializer

    def get_queryset(self):
        req: Any = self.request
        user = req.user
        params = getattr(req, 'query_params', getattr(req, 'GET', {}))
        tenant = getattr(req, 'tenant', None) or getattr(user, 'tenant', None)

        is_super = getattr(user, 'is_superuser', False) or getattr(user, 'role', '') == 'SUPER_ADMIN'

        if is_super:
            tenant_param = params.get('tenant_id')
            if tenant_param:
                qs = Student.objects.filter(tenant_id=tenant_param)
            elif tenant:
                qs = Student.objects.filter(tenant=tenant)
            else:
                qs = Student.objects.all()
        else:
            qs = Student.objects.filter(tenant=tenant)

        # For detail actions (retrieve, update, set_color, set_folders, etc.), return base queryset
        if self.action != 'list':
            return qs

        # ── 1. Folder & Archive/Hidden Scopes ──────────────────────────────
        folder = params.get('folder', 'all')
        include_archive = str(params.get('include_archive', 'false')).lower() == 'true'
        search_query = str(params.get('search', '')).strip()

        if folder == 'deleted' or folder == 'archive':
            qs = qs.filter(is_deleted=True)
        elif folder == 'hidden':
            qs = qs.filter(is_deleted=False, status_hidden=True)
        elif folder == 'except':
            qs = qs.filter(is_deleted=False).filter(Q(folder_ids=[]) | Q(folder_ids__isnull=True))
        elif folder != 'all':
            try:
                folder_uuid = uuid.UUID(str(folder).strip())
                qs = qs.filter(is_deleted=False, folder_ids__contains=[folder_uuid])
            except (ValueError, TypeError):
                qs = qs.none()
        else:
            if not include_archive and not search_query:
                qs = qs.filter(is_deleted=False)

        # ── 2. Search Filter ──────────────────────────────────────────────
        search_mode = params.get('search_mode', 'all')

        if search_query:
            if search_mode == 'id':
                qs = qs.filter(id__icontains=search_query)
            else:
                qs = qs.filter(
                    Q(id__icontains=search_query) |
                    Q(full_name__icontains=search_query) |
                    Q(korean_name__icontains=search_query) |
                    Q(passport__icontains=search_query) |
                    Q(phone1__icontains=search_query) |
                    Q(phone2__icontains=search_query) |
                    Q(father_name__icontains=search_query) |
                    Q(mother_name__icontains=search_query) |
                    Q(university_1__icontains=search_query)
                )

        # ── 3. Multi-Select Criteria Filters ──────────────────────────────
        getlist = getattr(params, 'getlist', None)
        tariffs = (getlist('tariff') if getlist else None) or (params.get('tariff', '').split(',') if params.get('tariff') else [])
        if tariffs and tariffs[0]:
            if 'NO_TARIFF' in tariffs or 'No Tariff' in tariffs:
                clean_tariffs = [t for t in tariffs if t not in ('NO_TARIFF', 'No Tariff')]
                qs = qs.filter(Q(tariff__in=clean_tariffs) | Q(tariff__isnull=True) | Q(tariff=''))
            else:
                qs = qs.filter(tariff__in=tariffs)

        levels = (getlist('level') if getlist else None) or (params.get('level', '').split(',') if params.get('level') else [])
        if levels and levels[0]:
            if 'NO_LEVEL' in levels or 'No Level' in levels:
                clean_levels = [l for l in levels if l not in ('NO_LEVEL', 'No Level')]
                qs = qs.filter(Q(level__in=clean_levels) | Q(level2__in=clean_levels) | Q(level__isnull=True) | Q(level=''))
            else:
                qs = qs.filter(Q(level__in=levels) | Q(level2__in=levels))

        groups = (getlist('group') if getlist else None) or (params.get('group', '').split(',') if params.get('group') else [])
        if groups and groups[0]:
            if 'NO_GROUP' in groups or 'No Group' in groups:
                clean_groups = [g for g in groups if g not in ('NO_GROUP', 'No Group')]
                qs = qs.filter(Q(student_group__in=clean_groups) | Q(student_group__isnull=True) | Q(student_group=''))
            else:
                qs = qs.filter(student_group__in=groups)

        certs = (getlist('cert') if getlist else None) or (params.get('cert', '').split(',') if params.get('cert') else [])
        if certs and certs[0]:
            cert_q = Q()
            for c in certs:
                if c == 'NO CERTIFICATE':
                    cert_q |= Q(language_certificate='NO CERTIFICATE') | Q(language_certificate__isnull=True) | Q(language_certificate='')
                else:
                    cert_q |= Q(language_certificate=c) | Q(language_certificate_2=c) | Q(language_certificate_3=c)
            qs = qs.filter(cert_q)

        scores = (getlist('score') if getlist else None) or (params.get('score', '').split(',') if params.get('score') else [])
        if scores and scores[0]:
            score_q = Q()
            for s in scores:
                score_q |= Q(certificate_score__iexact=s) | Q(certificate_score_2__iexact=s) | Q(certificate_score_3__iexact=s)
            qs = qs.filter(score_q)

        tags = (getlist('tag') if getlist else None) or (params.get('tag', '').split(',') if params.get('tag') else [])
        if tags and tags[0]:
            tag_q = Q()
            for t in tags:
                tag_q |= Q(task_tags__contains=t)
            qs = qs.filter(tag_q)

        leads = (getlist('lead_by') if getlist else None) or (params.get('lead_by', '').split(',') if params.get('lead_by') else [])
        if leads and leads[0]:
            if 'NO_LEADBY' in leads or 'No Lead by' in leads:
                clean_leads = [l for l in leads if l not in ('NO_LEADBY', 'No Lead by')]
                qs = qs.filter(Q(lead_by__in=clean_leads) | Q(lead_by__isnull=True) | Q(lead_by=''))
            else:
                qs = qs.filter(lead_by__in=leads)

        office = params.get('office')
        if office:
            qs = qs.filter(office=office)

        sort_by = params.get('sort_by', 'id')
        sort_order = params.get('sort_order', 'asc')

        if sort_by != 'id':
            order_prefix = '-' if sort_order == 'desc' else ''
            qs = qs.order_by(f"{order_prefix}{sort_by}")

        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        params = getattr(request, 'query_params', getattr(request, 'GET', {}))
        sort_by = params.get('sort_by', 'id')
        sort_order = params.get('sort_order', 'asc')

        if sort_by == 'id':
            student_list = list(queryset)
            student_list.sort(
                key=lambda s: alphanumeric_key(s.id),
                reverse=(sort_order == 'desc')
            )
            page = self.paginate_queryset(student_list)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(student_list, many=True)
            return Response(serializer.data)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        req: Any = self.request
        user = req.user
        tenant = getattr(req, 'tenant', None) or getattr(user, 'tenant', None)
        serializer.save(
            tenant=tenant,
            created_by=user
        )

    # ── Actions: Archive / Restore / Permanent Delete ──────────────────────
    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        student = self.get_object()
        archive_student(student, request.user)
        return Response({'status': 'Student archived'})

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        student = self.get_object()
        restore_student(student, request.user)
        return Response({'status': 'Student restored'})

    @action(detail=True, methods=['delete'], permission_classes=[IsTenantHeadManager])
    def permanent_delete(self, request, pk=None):
        student = self.get_object()
        permanent_delete_student(student, request.user)
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ── Actions: Color / Folder Assignment / Custom Tags ───────────────────
    @action(detail=True, methods=['post'])
    def set_color(self, request, pk=None):
        student = self.get_object()
        serializer = StudentSetColorSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        if 'row_color' in serializer.validated_data:
            student.row_color = serializer.validated_data['row_color']
        if 'status_row_color' in serializer.validated_data:
            student.status_row_color = serializer.validated_data['status_row_color']
        student.save(update_fields=['row_color', 'status_row_color', 'updated_at'])
        return Response({
            'status': 'Color updated',
            'row_color': student.row_color,
            'status_row_color': student.status_row_color
        })

    @action(detail=True, methods=['post'])
    def set_folders(self, request, pk=None):
        student = self.get_object()
        serializer = StudentSetFoldersSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        raw_ids = serializer.validated_data.get('folder_ids', [])
        clean_ids = []
        for fid in raw_ids:
            try:
                clean_ids.append(uuid.UUID(str(fid)))
            except (ValueError, TypeError):
                pass
        student.folder_ids = clean_ids
        student.save(update_fields=['folder_ids', 'updated_at'])
        return Response({
            'status': 'Folders updated',
            'folder_ids': [str(fid) for fid in student.folder_ids]
        })

    @action(detail=True, methods=['post'])
    def toggle_tag(self, request, pk=None):
        student = self.get_object()
        tag_name = request.data.get('tag')
        if tag_name:
            current_tags = list(student.task_tags or [])
            if tag_name in current_tags:
                current_tags.remove(tag_name)
            else:
                current_tags.append(tag_name)
            student.task_tags = current_tags
            student.save(update_fields=['task_tags', 'updated_at'])
        return Response({
            'status': 'Tag toggled',
            'task_tags': student.task_tags
        })

    @action(detail=True, methods=['post'])
    def clear_all(self, request, pk=None):
        student = self.get_object()
        student.row_color = None
        student.status_row_color = None
        student.task_tags = []
        student.save(update_fields=['row_color', 'status_row_color', 'task_tags', 'updated_at'])
        return Response({
            'status': 'Cleared',
            'row_color': None,
            'task_tags': []
        })


class FolderViewSet(viewsets.ModelViewSet):
    """Student Folders CRUD ViewSet."""
    serializer_class = FolderSerializer
    permission_classes = [IsTenantUser]
    pagination_class = None

    def get_queryset(self):
        req: Any = self.request
        user = req.user
        tenant = getattr(req, 'tenant', None) or getattr(user, 'tenant', None)
        if not tenant:
            from apps.tenants.models import Tenant
            tenant = Tenant.objects.first()
        if tenant:
            if not Folder.objects.filter(tenant=tenant, name__iexact='KDB').exists():
                Folder.objects.create(tenant=tenant, name='KDB')
            return Folder.objects.filter(tenant=tenant).order_by('name')
        return Folder.objects.all().order_by('name')

    def perform_create(self, serializer):
        req: Any = self.request
        user = req.user
        tenant = getattr(req, 'tenant', None) or getattr(user, 'tenant', None)
        serializer.save(tenant=tenant, created_by=user)

    @action(detail=True, methods=['post'], url_path='add-students')
    def add_students(self, request, pk=None):
        folder = self.get_object()
        student_ids = request.data.get('student_ids', [])
        if not student_ids:
            return Response({'status': 'No students provided', 'added_count': 0})

        folder_uuid = folder.id
        folder_str = str(folder_uuid)
        students = Student.objects.filter(id__in=student_ids)
        updated = 0
        for s in students:
            curr = list(s.folder_ids or [])
            curr_strs = [str(x) for x in curr]
            if folder_str not in curr_strs:
                curr.append(folder_uuid)
                s.folder_ids = curr
                s.save(update_fields=['folder_ids', 'updated_at'])
                updated += 1
        return Response({'status': 'Students added', 'added_count': updated})

    @action(detail=True, methods=['post'], url_path='sync-students')
    def sync_students(self, request, pk=None):
        folder = self.get_object()
        student_ids = set(request.data.get('student_ids', []))
        folder_uuid = folder.id
        req: Any = self.request
        tenant = getattr(req, 'tenant', None) or getattr(req.user, 'tenant', None)
        qs = Student.objects.filter(tenant=tenant) if tenant else Student.objects.all()

        for s in qs:
            curr = list(s.folder_ids or [])
            has_folder = folder_uuid in curr
            should_have = s.id in student_ids
            if should_have and not has_folder:
                curr.append(folder_uuid)
                s.folder_ids = curr
                s.save(update_fields=['folder_ids', 'updated_at'])
            elif not should_have and has_folder:
                curr = [fid for fid in curr if fid != folder_uuid]
                s.folder_ids = curr
                s.save(update_fields=['folder_ids', 'updated_at'])
        return Response({'status': 'Folder students synced'})


class StudentOptionsViewSet(viewsets.ViewSet):
    """Returns combined option lists for filters, forms, and settings."""
    permission_classes = [IsTenantUser]

    def list(self, request):
        user = request.user
        tenant = getattr(request, 'tenant', None) or getattr(user, 'tenant', None)
        if tenant:
            if not Folder.objects.filter(tenant=tenant, name__iexact='KDB').exists():
                Folder.objects.create(tenant=tenant, name='KDB')

        tariffs = list(TariffOption.objects.filter(tenant=tenant).values('name', 'price')) if tenant else []
        levels = list(EducationLevelOption.objects.filter(tenant=tenant).values_list('name', flat=True)) if tenant else []
        groups = list(StudentGroupOption.objects.filter(tenant=tenant).values_list('name', flat=True)) if tenant else []
        leads = list(LeadSourceOption.objects.filter(tenant=tenant).values_list('name', flat=True)) if tenant else []
        coordinators = list(CoordinatorOption.objects.filter(tenant=tenant).values_list('name', flat=True)) if tenant else []
        universities = list(UniversityOption.objects.values_list('name', flat=True).order_by('name'))
        folders_qs = Folder.objects.filter(tenant=tenant).order_by('name') if tenant else Folder.objects.all().order_by('name')

        all_count = Student.objects.filter(tenant=tenant, is_deleted=False).count() if tenant else Student.objects.filter(is_deleted=False).count()
        except_count = Student.objects.filter(tenant=tenant, is_deleted=False).filter(Q(folder_ids=[]) | Q(folder_ids__isnull=True)).count() if tenant else Student.objects.filter(is_deleted=False).filter(Q(folder_ids=[]) | Q(folder_ids__isnull=True)).count()
        deleted_count = Student.objects.filter(tenant=tenant, is_deleted=True).count() if tenant else Student.objects.filter(is_deleted=True).count()
        hidden_count = Student.objects.filter(tenant=tenant, is_deleted=False, status_hidden=True).count() if tenant else Student.objects.filter(is_deleted=False, status_hidden=True).count()

        folder_counts = {
            'all': all_count,
            'except': except_count,
            'deleted': deleted_count,
            'archive': deleted_count,
            'hidden': hidden_count,
        }
        for f in folders_qs:
            f_qs = Student.objects.filter(is_deleted=False, folder_ids__contains=[f.id])
            if tenant:
                f_qs = f_qs.filter(tenant=tenant)
            folder_counts[str(f.id)] = f_qs.count()

        folders = [
            {
                'id': str(f.id),
                'name': f.name,
                'student_count': folder_counts.get(str(f.id), 0)
            }
            for f in folders_qs
        ]

        tags_qs = TagOption.objects.filter(tenant=tenant).order_by('name') if tenant else TagOption.objects.all().order_by('name')
        if tenant and not tags_qs.exists():
            for t_name, t_icon in DEFAULT_TAGS_DATA:
                TagOption.objects.get_or_create(tenant=tenant, name=t_name, defaults={'icon': t_icon})
            tags_qs = TagOption.objects.filter(tenant=tenant).order_by('name')

        tags = [
            {
                'id': str(t.id),
                'name': t.name,
                'icon': t.icon
            }
            for t in tags_qs
        ]

        return Response({
            'tariffs': tariffs,
            'levels': levels,
            'groups': groups,
            'leads': leads,
            'coordinators': coordinators,
            'universities': universities,
            'folders': folders,
            'folder_counts': folder_counts,
            'offices': ['ANDIJON OFFIS', 'TOSHKENT OFFIS'],
            'tags': tags
        })


class StudentExportView(APIView):
    """
    Exports filtered or full student roster to Excel XLSX spreadsheet.
    """
    permission_classes = [IsTenantUser]

    def get(self, request):
        user = request.user
        tenant = getattr(request, 'tenant', None) or getattr(user, 'tenant', None)
        params = getattr(request, 'query_params', getattr(request, 'GET', {}))
        is_super = getattr(user, 'is_superuser', False) or getattr(user, 'role', '') == 'SUPER_ADMIN'

        if is_super:
            tenant_param = params.get('tenant_id')
            if tenant_param:
                qs = Student.objects.filter(tenant_id=tenant_param)
            elif tenant:
                qs = Student.objects.filter(tenant=tenant)
            else:
                qs = Student.objects.all()
        else:
            qs = Student.objects.filter(tenant=tenant)

        # Folder / Archive scope
        folder = params.get('folder', 'all')
        include_archive = str(params.get('include_archive', 'false')).lower() == 'true'

        if folder == 'deleted' or folder == 'archive':
            qs = qs.filter(is_deleted=True)
        elif folder == 'hidden':
            qs = qs.filter(is_deleted=False, status_hidden=True)
        elif folder == 'except':
            qs = qs.filter(is_deleted=False).filter(Q(folder_ids=[]) | Q(folder_ids__isnull=True))
        elif folder != 'all':
            try:
                folder_uuid = uuid.UUID(str(folder).strip())
                qs = qs.filter(is_deleted=False, folder_ids__contains=[folder_uuid])
            except (ValueError, TypeError):
                qs = qs.none()
        else:
            if not include_archive:
                qs = qs.filter(is_deleted=False)

        # Search filter
        search_query = str(params.get('search', '')).strip()
        search_mode = params.get('search_mode', 'all')
        if search_query:
            if search_mode == 'id':
                qs = qs.filter(id__icontains=search_query)
            else:
                qs = qs.filter(
                    Q(id__icontains=search_query) |
                    Q(full_name__icontains=search_query) |
                    Q(korean_name__icontains=search_query) |
                    Q(passport__icontains=search_query) |
                    Q(phone1__icontains=search_query) |
                    Q(phone2__icontains=search_query)
                )

        # Multi-select filters
        getlist = getattr(params, 'getlist', None)
        tariffs = (getlist('tariff') if getlist else None) or (params.get('tariff', '').split(',') if params.get('tariff') else [])
        if tariffs and tariffs[0]:
            if 'No Tariff' in tariffs or 'NO_TARIFF' in tariffs:
                clean_tariffs = [t for t in tariffs if t not in ('NO_TARIFF', 'No Tariff')]
                qs = qs.filter(Q(tariff__in=clean_tariffs) | Q(tariff__isnull=True) | Q(tariff=''))
            else:
                qs = qs.filter(tariff__in=tariffs)

        levels = (getlist('level') if getlist else None) or (params.get('level', '').split(',') if params.get('level') else [])
        if levels and levels[0]:
            qs = qs.filter(Q(level__in=levels) | Q(level2__in=levels))

        students_list = sorted(list(qs), key=lambda s: alphanumeric_key(s.id))

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title="Students Roster")
        else:
            ws.title = "Students Roster"

        headers = [
            "ID", "Full Name", "Korean Name", "Phone 1", "Phone 2", "Tariff",
            "Level", "Language Certificate", "Certificate Score", "Language Cert 2", "Cert 2 Score",
            "University 1", "University 2", "University 3", "University 4", "University 5",
            "Student Group", "Lead By", "Coordinator", "CoA", "Invoice", "Embassy"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for s in students_list:
            ws.append([
                s.id,
                s.full_name,
                s.korean_name or '',
                s.phone1 or '',
                s.phone2 or '',
                s.tariff or 'NO TARIFF',
                s.level or '',
                s.language_certificate or '',
                s.certificate_score or '',
                s.language_certificate_2 or '',
                s.certificate_score_2 or '',
                s.university_1 or '',
                s.university_2 or '',
                s.university_3 or '',
                s.university_4 or '',
                s.university_5 or '',
                s.student_group or '',
                s.lead_by or '',
                s.coordinator or '',
                s.coa or 'NOT TAKEN',
                s.invoice or 'NOT TAKEN',
                s.embassy or 'NOT TAKEN'
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            first_cell = col[0]
            if first_cell.column is not None:
                col_letter = get_column_letter(int(first_cell.column))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="salom_crm_students.xlsx"'
        return response


class BaseOptionViewSet(viewsets.ModelViewSet):
    model_class: Any = None
    permission_classes = [IsTenantUser]
    pagination_class = None

    def get_queryset(self):
        req: Any = self.request
        user = req.user
        tenant = getattr(req, 'tenant', None) or getattr(user, 'tenant', None)
        if not tenant:
            from apps.tenants.models import Tenant
            tenant = Tenant.objects.first()
        if not tenant or self.model_class is None:
            return self.model_class.objects.none() if self.model_class else Student.objects.none()
        return self.model_class.objects.filter(tenant=tenant).order_by('name')

    def perform_create(self, serializer):
        req: Any = self.request
        user = req.user
        tenant = getattr(req, 'tenant', None) or getattr(user, 'tenant', None)
        if not tenant:
            from apps.tenants.models import Tenant
            tenant = Tenant.objects.first()
        serializer.save(tenant=tenant)


class TariffOptionViewSet(BaseOptionViewSet):
    model_class = TariffOption
    serializer_class = TariffOptionSerializer


class EducationLevelOptionViewSet(BaseOptionViewSet):
    model_class = EducationLevelOption
    serializer_class = EducationLevelOptionSerializer


class StudentGroupOptionViewSet(BaseOptionViewSet):
    model_class = StudentGroupOption
    serializer_class = StudentGroupOptionSerializer


class LeadSourceOptionViewSet(BaseOptionViewSet):
    model_class = LeadSourceOption
    serializer_class = LeadSourceOptionSerializer


class CoordinatorOptionViewSet(BaseOptionViewSet):
    model_class = CoordinatorOption
    serializer_class = CoordinatorOptionSerializer


class UniversityOptionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsTenantUser]
    pagination_class = None
    serializer_class = UniversityOptionSerializer
    queryset = UniversityOption.objects.all().order_by('name')


class UniversityStatusOptionViewSet(BaseOptionViewSet):
    model_class = UniversityStatusOption
    serializer_class = UniversityStatusOptionSerializer


class TagOptionViewSet(BaseOptionViewSet):
    model_class = TagOption
    serializer_class = TagOptionSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        req: Any = self.request
        tenant = getattr(req, 'tenant', None) or getattr(req.user, 'tenant', None)
        if tenant and not qs.exists():
            for t_name, t_icon in DEFAULT_TAGS_DATA:
                TagOption.objects.get_or_create(tenant=tenant, name=t_name, defaults={'icon': t_icon})
            qs = TagOption.objects.filter(tenant=tenant).order_by('name')
        return qs


class SchoolDirectoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsTenantUser]
    pagination_class = None
    serializer_class = SchoolDirectorySerializer
    queryset = SchoolDirectory.objects.all().order_by('name')

    @action(detail=False, methods=['post'], url_path='upsert')
    def upsert(self, request: Request):
        name = str(request.data.get('name', '')).strip()
        if not name:
            return Response({'error': 'School name is required'}, status=status.HTTP_400_BAD_REQUEST)

        address = request.data.get('address', None)
        website = request.data.get('website', None)
        phone = request.data.get('phone', None)
        email = request.data.get('email', None)

        school, created = SchoolDirectory.objects.update_or_create(
            name=name,
            defaults={
                'address': address if address is not None else '',
                'website': website if website is not None else '',
                'phone': phone if phone is not None else '',
                'email': email if email is not None else '',
            }
        )
        return Response(SchoolDirectorySerializer(school).data, status=status.HTTP_200_OK)


class MajorOptionViewSet(BaseOptionViewSet):
    model_class = MajorOption
    serializer_class = MajorOptionSerializer

    @action(detail=False, methods=['post'], url_path='upsert')
    def upsert(self, request: Request):
        tenant = getattr(request, 'tenant', None) or getattr(request.user, 'tenant', None)
        if not tenant:
            return Response({'error': 'Tenant required'}, status=status.HTTP_400_BAD_REQUEST)
        
        name = str(request.data.get('name', '')).strip().upper()
        if not name:
            return Response({'error': 'Major name is required'}, status=status.HTTP_400_BAD_REQUEST)

        major, created = MajorOption.objects.get_or_create(
            tenant=tenant,
            name=name,
            defaults={'created_by': request.user if request.user.is_authenticated else None}
        )
        return Response(MajorOptionSerializer(major).data, status=status.HTTP_200_OK)


class ExtractDocumentView(APIView):
    """
    Enterprise in-memory document extraction and OCR service for student profiles.
    Zero permanent file storage - processes in RAM with strict size and concurrency limits.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request):
        from .ocr_service import process_document_ephemeral
        from .ocr_preprocessor import PreprocessError

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response(
                {'error': 'A document file (PDF, JPG, PNG, WEBP) must be uploaded via multipart/form-data.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Enforce maximum upload size (10 MB)
        if file_obj.size > 10 * 1024 * 1024:
            return Response(
                {'error': 'File size exceeds maximum allowed limit of 10MB.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Enforce valid document extensions & magic bytes
        filename = file_obj.name or ''
        ext = filename.lower().split('.')[-1] if '.' in filename else ''
        if ext not in ('pdf', 'jpg', 'jpeg', 'png', 'webp', 'bmp', 'tiff'):
            return Response(
                {'error': 'Unsupported file format. Supported formats: PDF, JPG, PNG, WEBP, BMP.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file_bytes = file_obj.read()
        if not file_bytes:
            return Response({'error': 'Uploaded file is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        # Debug mode enabled for staff/superusers with ?debug=true
        is_debug = request.user.is_staff and (request.query_params.get('debug') in ('1', 'true'))

        try:
            extracted_data = process_document_ephemeral(file_bytes, filename, debug=is_debug)

            # Check student ID for parent passport intelligence (15+ years older than student)
            student_id = request.data.get('student_id') or request.query_params.get('student_id')
            if student_id:
                try:
                    student_qs = Student.objects.filter(id=student_id)
                    if request.user and hasattr(request.user, 'tenant') and request.user.tenant:
                        student_qs = student_qs.filter(tenant=request.user.tenant)
                    student = student_qs.first()

                    if student and student.birthday and extracted_data.get('document_type') in ('PASSPORT', 'ID_CARD'):
                        extracted_dob = extracted_data.get('fields', {}).get('DATE_OF_BIRTH')
                        if extracted_dob:
                            doc_year = None
                            student_year = None

                            # Parse 4-digit years from ISO YYYY-MM-DD or date strings
                            m_doc = re.search(r'\b(19\d\d|20\d\d)\b', str(extracted_dob))
                            m_stu = re.search(r'\b(19\d\d|20\d\d)\b', str(student.birthday))

                            if m_doc and m_stu:
                                doc_year = int(m_doc.group(1))
                                student_year = int(m_stu.group(1))
                                age_diff = student_year - doc_year
                            else:
                                age_diff = 20 if str(extracted_dob) < str(student.birthday) else 0

                            # If the passport holder is 15+ years older than the student -> Parent Passport!
                            if age_diff >= 15:
                                sex = str(extracted_data.get('fields', {}).get('SEX', '')).upper()
                                full_name = extracted_data.get('fields', {}).get('FULL_NAME', '')

                                # Infer gender if sex field wasn't clearly detected by OCR
                                is_female = (sex in ('F', 'FEMALE', 'AYOL', 'ЖЕН', 'ЖЕНСКИЙ'))
                                if not is_female and not (sex in ('M', 'MALE', 'ERKAK', 'МУЖ', 'МУЖСКОЙ')):
                                    if any(q in full_name.upper() for q in ['QIZI', 'KIZI', 'OVNA', 'EVNA', 'KYZY']):
                                        is_female = True

                                parent_field = 'MOTHER_FULLNAME' if is_female else 'FATHER_FULLNAME'
                                parent_doc_type = 'MOTHER_PASSPORT' if is_female else 'FATHER_PASSPORT'

                                extracted_data['is_parent_passport'] = True
                                extracted_data['parent_type'] = 'MOTHER' if is_female else 'FATHER'
                                extracted_data['document_type'] = parent_doc_type

                                # ISOLATE FIELDS: Only keep parent name, delete student's personal identity fields
                                # so parent's passport number, DOB, and sex are NOT written to the student's profile!
                                extracted_data['fields'] = {
                                    parent_field: full_name
                                }
                                extracted_data['field_details'] = {
                                    parent_field: {
                                        'value': full_name,
                                        'confidence': 0.98,
                                        'validated': True,
                                        'source': 'PARENT_PASSPORT'
                                    }
                                }
                except Exception as e:
                    logger.warning(f"Error evaluating parent passport intelligence: {e}")

            return Response(extracted_data, status=status.HTTP_200_OK)
        except PreprocessError as pe:
            return Response({'error': str(pe)}, status=status.HTTP_400_BAD_REQUEST)
        except TimeoutError as te:
            return Response({'error': str(te)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except Exception as e:
            return Response({'error': f'Failed to process document: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VisaCheckView(APIView):
    """
    Checks visa application status directly from visa.go.kr.
    Supports Embassy (gb03), E-Visa (gb01), and Regional (gb02).
    """
    permission_classes = [IsTenantUser]

    def post(self, request: Request) -> Response:
        passport = request.data.get('passport', '').strip()
        full_name = (request.data.get('full_name') or request.data.get('name') or '').strip()
        birth_date = (request.data.get('birth_date') or request.data.get('dob') or '').strip()
        visa_type = request.data.get('visa_type', 'Embassy').strip()
        application_no = (request.data.get('application_no') or request.data.get('app_no') or '').strip()

        if not passport:
            return Response({'error': 'Passport number is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not full_name:
            return Response({'error': 'Full name is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not birth_date:
            return Response({'error': 'Date of birth is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if visa_type in ('E-Visa', 'Regional') and not application_no:
            return Response({'error': 'Application number is required for E-Visa/Regional visa.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from .visa_service import check_visa_direct
            from .models import VisaStudent
            from django.utils import timezone

            result = check_visa_direct(
                passport=passport,
                full_name=full_name,
                birth_date=birth_date,
                visa_type=visa_type,
                application_no=application_no
            )

            # Persist check results to VisaStudent database if student exists
            tenant = getattr(request.user, 'tenant', None)
            vs = VisaStudent.objects.filter(is_deleted=False)
            if tenant:
                vs = vs.filter(tenant=tenant)
            vs_obj = vs.filter(passport__iexact=passport).first()
            if vs_obj:
                if result.get('latest_status'):
                    vs_obj.status = result.get('latest_status').upper()
                if result.get('latest_date'):
                    vs_obj.application_date = result.get('latest_date')
                if result.get('entry_date'):
                    vs_obj.status_date = result.get('entry_date')
                vs_obj.rejection_reason = result.get('rejection_reason') or ''
                vs_obj.pdf_url = result.get('pdf_url') or ''
                vs_obj.api_response = result
                vs_obj.last_checked = timezone.now()
                vs_obj.save()

            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': f'visa.go.kr checking failed: {str(e)}'},
                status=status.HTTP_502_BAD_GATEWAY
            )


class VisaDownloadPdfView(APIView):
    """
    Downloads official visa certificate PDF from visa.go.kr.
    """
    permission_classes = [IsTenantUser]

    def post(self, request: Request) -> HttpResponse | Response:
        passport = request.data.get('passport', '').strip()
        full_name = (request.data.get('full_name') or request.data.get('name') or '').strip()
        birth_date = (request.data.get('birth_date') or request.data.get('dob') or '').strip()
        visa_type = request.data.get('visa_type', 'Embassy').strip()
        application_no = (request.data.get('application_no') or request.data.get('app_no') or '').strip()
        pdf_url = request.data.get('pdf_url', '').strip()

        if not passport or not full_name or not birth_date:
            return Response({'error': 'Passport, full name, and birth date are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from .visa_service import download_visa_pdf
            pdf_bytes = download_visa_pdf(
                passport=passport,
                full_name=full_name,
                birth_date=birth_date,
                visa_type=visa_type,
                application_no=application_no,
                pdf_url=pdf_url
            )
            resp = HttpResponse(pdf_bytes, content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="visa_{passport.upper()}.pdf"'
            return resp
        except Exception as e:
            return Response({'error': f'Failed to download visa PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request: Request) -> HttpResponse | Response:
        passport = request.query_params.get('passport', '').strip()
        full_name = (request.query_params.get('full_name') or request.query_params.get('name') or '').strip()
        birth_date = (request.query_params.get('birth_date') or request.query_params.get('dob') or '').strip()
        visa_type = request.query_params.get('visa_type', 'Embassy').strip()
        application_no = (request.query_params.get('application_no') or request.query_params.get('app_no') or '').strip()
        pdf_url = request.query_params.get('pdf_url', '').strip()

        if not passport or not full_name or not birth_date:
            return Response({'error': 'Passport, full name, and birth date are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from .visa_service import download_visa_pdf
            pdf_bytes = download_visa_pdf(
                passport=passport,
                full_name=full_name,
                birth_date=birth_date,
                visa_type=visa_type,
                application_no=application_no,
                pdf_url=pdf_url
            )
            resp = HttpResponse(pdf_bytes, content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="visa_{passport.upper()}.pdf"'
            return resp
        except Exception as e:
            return Response({'error': f'Failed to download visa PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VisaStudentQuickSearchView(APIView):
    """
    Returns quick student suggestions for autofilling visa check form.
    """
    permission_classes = [IsTenantUser]

    def get(self, request: Request) -> Response:
        q = request.query_params.get('q', '').strip()
        if not q or len(q) < 2:
            return Response([])

        tenant = getattr(request.user, 'tenant', None)
        qs = Student.objects.filter(is_deleted=False)
        if tenant:
            qs = qs.filter(tenant=tenant)

        qs = qs.filter(
            Q(full_name__icontains=q) |
            Q(passport__icontains=q) |
            Q(id__icontains=q)
        )[:10]

        results = []
        for s in qs:
            results.append({
                'id': s.id,
                'full_name': s.full_name,
                'passport': s.passport or '',
                'birthday': s.birthday or '',
                'tariff': s.tariff or '',
                'university': s.university_1 or s.invoice_university or '',
                'coordinator': s.coordinator or '',
                'embassy_visa_status': s.embassy or ''
            })

        return Response(results)


class VisaStudentLookupView(APIView):
    """
    Looks up a passport in the MAIN Student CRM database.
    Used for instant auto-filling New Student form.
    """
    permission_classes = [IsTenantUser]

    def get(self, request: Request) -> Response:
        passport = request.query_params.get('passport', '').strip()
        if not passport:
            return Response({'found': False})

        tenant = getattr(request.user, 'tenant', None)
        qs = Student.objects.filter(is_deleted=False)
        if tenant:
            qs = qs.filter(tenant=tenant)

        student = qs.filter(passport__iexact=passport).first()
        if not student:
            return Response({'found': False})

        return Response({
            'found': True,
            'student': {
                'id': student.id or '',
                'full_name': student.full_name or '',
                'passport': student.passport or '',
                'birthday': student.birthday or '',
                'tariff': student.tariff or '',
                'university': student.university_1 or student.invoice_university or '',
                'coordinator': student.coordinator or '',
                'phone1': student.phone1 or '',
            }
        })


class VisaStudentListCreateView(APIView):
    """
    Dedicated endpoints for Visa Check database table (crm_visa_students).
    Completely isolated from main Student table.
    """
    permission_classes = [IsTenantUser]

    def get(self, request: Request) -> Response:
        from .models import VisaStudent
        from .serializers import VisaStudentSerializer

        tenant = getattr(request.user, 'tenant', None)
        qs = VisaStudent.objects.filter(is_deleted=False)
        if tenant:
            qs = qs.filter(tenant=tenant)

        # Search filter
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search) |
                Q(passport__icontains=search) |
                Q(student_id__icontains=search) |
                Q(university__icontains=search) |
                Q(tariff__icontains=search)
            )

        # Status filter
        status_filter = request.query_params.get('status', '').strip().lower()
        if status_filter == 'pending':
            qs = qs.exclude(
                Q(status__icontains='APPROV') |
                Q(status__icontains='VISA USED') |
                Q(status__icontains='REJECT') |
                Q(status__icontains='CANCEL') |
                Q(status__icontains='RETURN') |
                Q(status__icontains='EXPIRED')
            )
        elif status_filter == 'approved':
            qs = qs.filter(
                Q(status__icontains='APPROV') |
                Q(status__icontains='VISA USED')
            )
        elif status_filter == 'cancelled':
            qs = qs.filter(
                Q(status__icontains='REJECT') |
                Q(status__icontains='CANCEL') |
                Q(status__icontains='RETURN') |
                Q(status__icontains='EXPIRED')
            )

        # Sorting
        sort_by = request.query_params.get('sort_by', 'date')
        if sort_by == 'university':
            qs = qs.order_by('-pinned', 'university', '-created_at')
        elif sort_by == 'tariff':
            qs = qs.order_by('-pinned', 'tariff', '-created_at')
        elif sort_by == 'selected':
            qs = qs.order_by('-pinned', '-batch_selected', '-created_at')
        elif sort_by == 'statusDate':
            qs = qs.order_by('-pinned', '-status_date', '-created_at')
        else:
            qs = qs.order_by('-pinned', '-created_at')

        serializer = VisaStudentSerializer(qs, many=True)
        return Response({
            'count': qs.count(),
            'results': serializer.data
        })

    def post(self, request: Request) -> Response:
        from .models import VisaStudent
        from .serializers import VisaStudentSerializer

        tenant = getattr(request.user, 'tenant', None)
        passport = request.data.get('passport', '').strip().upper()
        if not passport:
            return Response({'error': 'Passport is required.'}, status=status.HTTP_400_BAD_REQUEST)

        full_name = request.data.get('full_name', '').strip().upper()
        if not full_name:
            return Response({'error': 'Full name is required.'}, status=status.HTTP_400_BAD_REQUEST)

        data = cast(dict[str, Any], request.data) if isinstance(request.data, dict) else {}
        s_id = str(data.get('student_id') or data.get('id') or '').strip().upper()

        visa_student, created = VisaStudent.objects.get_or_create(
            tenant=tenant,
            passport=passport,
            defaults={
                'student_id': s_id,
                'full_name': full_name,
                'birthday': str(data.get('birthday', '')).strip(),
                'visa_type': str(data.get('visa_type', 'Embassy')),
                'application_no': str(data.get('application_no', '')).strip().upper(),
                'tariff': str(data.get('tariff', '')),
                'university': str(data.get('university', '')),
                'coordinator': str(data.get('coordinator', '')),
                'b2b': str(data.get('b2b', '')),
                'flag': bool(data.get('flag', False)),
                'refund_application': bool(data.get('refund_application', False)),
            }
        )

        if not created:
            # Update existing
            for field in ('full_name', 'birthday', 'visa_type', 'application_no', 'tariff', 'university', 'coordinator', 'b2b', 'flag', 'refund_application'):
                if field in data:
                    setattr(visa_student, field, data[field])
            if s_id:
                visa_student.student_id = s_id
            visa_student.is_deleted = False
            visa_student.save()

        serializer = VisaStudentSerializer(visa_student)
        return Response(serializer.data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


class VisaStudentDetailView(APIView):
    """
    Retrieve, Patch individual management fields, or Delete a Visa Student.
    Does NOT affect the main CRM Student database.
    """
    permission_classes = [IsTenantUser]

    def _get_student(self, request: Request, passport: str):
        from .models import VisaStudent
        tenant = getattr(request.user, 'tenant', None)
        qs = VisaStudent.objects.filter(is_deleted=False)
        if tenant:
            qs = qs.filter(tenant=tenant)
        return qs.filter(Q(passport__iexact=passport) | Q(student_id__iexact=passport)).first()

    def get(self, request: Request, passport: str) -> Response:
        from .serializers import VisaStudentSerializer
        student = self._get_student(request, passport)
        if not student:
            return Response({'error': 'Student not found in Visa database.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(VisaStudentSerializer(student).data)

    def patch(self, request: Request, passport: str) -> Response:
        from .serializers import VisaStudentSerializer
        student = self._get_student(request, passport)
        if not student:
            return Response({'error': 'Student not found in Visa database.'}, status=status.HTTP_404_NOT_FOUND)

        data = cast(dict[str, Any], request.data) if isinstance(request.data, dict) else {}

        # Allow partial updating of any management or status fields
        for field in (
            'student_id', 'full_name', 'birthday', 'visa_type', 'application_no',
            'tariff', 'university', 'coordinator', 'b2b',
            'flag', 'refund_application', 'pinned', 'batch_selected',
            'status', 'rejection_reason', 'pdf_url', 'application_date', 'status_date'
        ):
            if field in data:
                val = data[field]
                if field in ('flag', 'refund_application', 'pinned', 'batch_selected'):
                    setattr(student, field, bool(val))
                elif field in ('tariff', 'university', 'coordinator', 'b2b'):
                    setattr(student, field, '' if val in ('none', 'None', None) else str(val))
                else:
                    setattr(student, field, val)

        if 'id' in data and 'student_id' not in data:
            student.student_id = str(data['id'])

        student.save()
        return Response(VisaStudentSerializer(student).data)

    def delete(self, request: Request, passport: str) -> Response:
        student = self._get_student(request, passport)
        if not student:
            return Response({'error': 'Student not found in Visa database.'}, status=status.HTTP_404_NOT_FOUND)

        # Hard delete from visa check database table
        student.delete()
        return Response({'message': f'Student {passport} removed from Visa database.'}, status=status.HTTP_200_OK)


class VisaStudentBulkDeleteView(APIView):
    """
    Deletes multiple students from the Visa database table.
    """
    permission_classes = [IsTenantUser]

    def post(self, request: Request) -> Response:
        from .models import VisaStudent
        tenant = getattr(request.user, 'tenant', None)
        passports = request.data.get('passports', [])
        if not passports or not isinstance(passports, list):
            return Response({'error': 'Passports list is required.'}, status=status.HTTP_400_BAD_REQUEST)

        qs = VisaStudent.objects.all()
        if tenant:
            qs = qs.filter(tenant=tenant)

        upper_passports = [str(p).strip().upper() for p in passports]
        deleted_count, _ = qs.filter(passport__in=upper_passports).delete()
        return Response({'deleted_count': deleted_count, 'message': f'{deleted_count} student(s) deleted.'})


class VisaOptionsView(APIView):
    """
    Returns dropdown choices for Visa Check Management (Tariffs, Universities, Coordinators, B2B).
    """
    permission_classes = [IsTenantUser]

    def get(self, request: Request) -> Response:
        from .models import TariffOption, UniversityOption, CoordinatorOption, B2BOption

        tenant = getattr(request.user, 'tenant', None)

        def get_names(model_cls, defaults=None):
            qs = model_cls.objects.all()
            if tenant and hasattr(model_cls, 'tenant'):
                qs = qs.filter(tenant=tenant)
            names = list(qs.values_list('name', flat=True))
            if not names and defaults:
                names = defaults
            return [{'name': n} for n in names if n]

        tariffs = get_names(TariffOption, ['STANDART', 'PREMIUM', 'VISA PLUS', 'E-VISA TIL SERTIFIKATSIZ', 'E-VISA TIL SERTIFIKATLI', 'REGIONAL', 'ZERO RISK'])
        universities = get_names(UniversityOption, ['Baekseok University', 'Jeonju University', 'Anyang University', 'Hoseo University', 'Seoul National University'])
        coordinators = get_names(CoordinatorOption, ['Coordinator 1', 'Coordinator 2'])
        b2b = get_names(B2BOption, ['iTOP EDU', 'Global Edu', 'Direct Partner'])

        return Response({
            'tariffs': tariffs,
            'universities': universities,
            'coordinators': coordinators,
            'b2b': b2b
        })



