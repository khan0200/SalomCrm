"""
Unit tests for Excel Fill service (apps.students.excel_fill_service).
Validates CRM_FIELDS dictionary, semantic pattern matching, formatting rules,
and filled Excel template generation.
"""

import io
import unittest
import openpyxl

from apps.students.excel_fill_service import (
    CRM_FIELDS,
    SEMANTIC_PATTERNS,
    DATE_FIELDS,
    match_column_field,
    format_cell_value,
    generate_filled_excel,
)


class TestExcelFillFieldsAndMatching(unittest.TestCase):
    def test_crm_fields_contain_all_requested_fields(self):
        keys = {f['key'] for f in CRM_FIELDS}
        # Educational background and school
        self.assertIn('educational_background', keys)
        self.assertIn('date_of_entry', keys)
        self.assertIn('date_of_graduation', keys)
        self.assertIn('graduation_expected', keys)
        self.assertIn('degree_no', keys)
        self.assertIn('gpa_system', keys)
        self.assertIn('school_address', keys)
        self.assertIn('school_phone', keys)
        self.assertIn('school_email', keys)
        self.assertIn('school_website', keys)
        self.assertIn('level2', keys)

        # Dates and certificates
        self.assertIn('today_date', keys)
        self.assertIn('certificate_test_date', keys)
        self.assertIn('language_certificate_2', keys)
        self.assertIn('certificate_score_2', keys)
        self.assertIn('certificate_2_test_date', keys)
        self.assertIn('certificate_2_valid_date', keys)
        self.assertIn('language_certificate_3', keys)
        self.assertIn('certificate_score_3', keys)
        self.assertIn('certificate_3_test_date', keys)
        self.assertIn('certificate_3_valid_date', keys)

        # University choices
        self.assertIn('university_1', keys)
        self.assertIn('university_1_major', keys)
        self.assertIn('university_2', keys)
        self.assertIn('university_3', keys)

        # ID & Management
        self.assertIn('id', keys)
        self.assertIn('tariff', keys)
        self.assertIn('student_group', keys)
        self.assertIn('coordinator', keys)

    def test_semantic_matching_entry_and_graduation_dates(self):
        # Entry dates
        for header in [
            'Date of Entry', 'Entry Date', 'Data of entry', 'Admission Date',
            'Entrance Date', '입학일', '입학일자', '입학년월일', 'kirish sanasi'
        ]:
            field, conf = match_column_field(header)
            self.assertEqual(field, 'date_of_entry', f"Failed for header {header!r}")
            self.assertGreaterEqual(conf, 0.8)

        # Graduation dates
        for header in [
            'Date of Graduation', 'Graduation Date', 'Data of graduation',
            'Graduated Date', '졸업일', '졸업일자', '졸업년월일', 'bitirgan sana'
        ]:
            field, conf = match_column_field(header)
            self.assertEqual(field, 'date_of_graduation', f"Failed for header {header!r}")
            self.assertGreaterEqual(conf, 0.8)

        # Graduation expected
        for header in ['Graduation Expected', 'Expected Graduation', '졸업예정']:
            field, conf = match_column_field(header)
            self.assertEqual(field, 'graduation_expected', f"Failed for header {header!r}")

    def test_semantic_matching_educational_background(self):
        for header in [
            'Educational Background', 'Education Background', 'Education Level',
            '최종학력 구분', '학력사항', '학력구분'
        ]:
            field, conf = match_column_field(header)
            self.assertEqual(field, 'educational_background', f"Failed for header {header!r}")

    def test_semantic_matching_degree_no_and_school_details(self):
        self.assertEqual(match_column_field('Degree No')[0], 'degree_no')
        self.assertEqual(match_column_field('Diploma Number')[0], 'degree_no')
        self.assertEqual(match_column_field('학위번호')[0], 'degree_no')
        self.assertEqual(match_column_field('School Address')[0], 'school_address')
        self.assertEqual(match_column_field('학교 주소')[0], 'school_address')
        self.assertEqual(match_column_field('School Phone')[0], 'school_phone')
        self.assertEqual(match_column_field('학교 연락처')[0], 'school_phone')

    def test_semantic_matching_target_university_and_today(self):
        self.assertEqual(match_column_field('Applying University')[0], 'university_1')
        self.assertEqual(match_column_field('지원대학')[0], 'university_1')
        self.assertEqual(match_column_field('Application Date')[0], 'today_date')
        self.assertEqual(match_column_field('신청일자')[0], 'today_date')
        self.assertEqual(match_column_field('Test Date')[0], 'certificate_test_date')
        self.assertEqual(match_column_field('응시일')[0], 'certificate_test_date')


class TestExcelFormattingAndGeneration(unittest.TestCase):
    def test_date_formatting_for_all_date_fields(self):
        self.assertIn('date_of_entry', DATE_FIELDS)
        self.assertIn('date_of_graduation', DATE_FIELDS)
        self.assertIn('certificate_test_date', DATE_FIELDS)
        self.assertIn('today_date', DATE_FIELDS)

        raw = '2020-09-02'
        self.assertEqual(
            format_cell_value(raw, 'date_of_entry', {'dateFormat': 'YYYY.MM.DD'}),
            '2020.09.02'
        )
        self.assertEqual(
            format_cell_value(raw, 'date_of_graduation', {'dateFormat': 'YYYYMMDD'}),
            '20200902'
        )
        self.assertEqual(
            format_cell_value(raw, 'certificate_test_date', {'dateFormat': 'DD.MM.YYYY'}),
            '02.09.2020'
        )

    def test_graduation_expected_formatting(self):
        self.assertEqual(
            format_cell_value(True, 'graduation_expected', {'boolFormat': 'Yes/No'}),
            'Yes'
        )
        self.assertEqual(
            format_cell_value(False, 'graduation_expected', {'boolFormat': 'Yes/No'}),
            'No'
        )
        self.assertEqual(
            format_cell_value('true', 'graduation_expected', {'boolFormat': '졸업예정/졸업'}),
            '졸업예정'
        )
        self.assertEqual(
            format_cell_value('false', 'graduation_expected', {'boolFormat': '졸업예정/졸업'}),
            '졸업'
        )

    def test_generate_filled_excel_with_new_fields(self):
        # Build in-memory template workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applicants"
        ws.cell(1, 1, "No")
        ws.cell(1, 2, "Full Name")
        ws.cell(1, 3, "Educational Background")
        ws.cell(1, 4, "Date of Entry")
        ws.cell(1, 5, "Date of Graduation")
        ws.cell(1, 6, "Degree No")

        buf = io.BytesIO()
        wb.save(buf)
        template_bytes = buf.getvalue()

        mappings = [
            {"col_idx": 1, "field": "_sequence_no"},
            {"col_idx": 2, "field": "full_name"},
            {"col_idx": 3, "field": "educational_background"},
            {"col_idx": 4, "field": "date_of_entry", "format_rules": {"dateFormat": "YYYY.MM.DD"}},
            {"col_idx": 5, "field": "date_of_graduation", "format_rules": {"dateFormat": "YYYY.MM.DD"}},
            {"col_idx": 6, "field": "degree_no"},
        ]

        students = [
            {
                "id": "UB101",
                "full_name": "ALIJON VALIYEV",
                "educational_background": "BACHELOR",
                "date_of_entry": "2020-09-02",
                "date_of_graduation": "2024-06-25",
                "degree_no": "B1234567",
            }
        ]

        output = generate_filled_excel(
            file_bytes=template_bytes,
            sheet_name="Applicants",
            column_mappings=mappings,
            students_data=students,
            fill_mode="append",
        )

        filled_wb = openpyxl.load_workbook(output)
        filled_ws = filled_wb["Applicants"]
        # Row 1 is header, Row 2 is filled student
        self.assertEqual(filled_ws.cell(2, 1).value, 1)
        self.assertEqual(filled_ws.cell(2, 2).value, "ALIJON VALIYEV")
        self.assertEqual(filled_ws.cell(2, 3).value, "BACHELOR")
        self.assertEqual(filled_ws.cell(2, 4).value, "2020.09.02")
        self.assertEqual(filled_ws.cell(2, 5).value, "2024.06.25")
        self.assertEqual(filled_ws.cell(2, 6).value, "B1234567")


if __name__ == '__main__':
    unittest.main()
